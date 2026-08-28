#!/usr/bin/env python3
"""
jarvis_rag_poc.py  —  Transparent RAG for Rick's fleet (Smedley)
================================================================
Pipeline (all already-live services, no black boxes):
  - Embeddings : LM Studio @127.0.0.1:1234  model text-embedding-nomic-embed-text-v1.5 (768d)
  - Answers    : LM Studio @127.0.0.1:1234  model from rag_config (DEFAULT_MODEL)
  - Vectors    : Qdrant @192.168.0.25:6333  collection 'jarvis_kb' (768d, Cosine)

Extraction by type:
  .doc/.docx/.rtf/.html -> macOS textutil
  .txt/.md/.log         -> read as text
  .pdf                  -> sidecar-first: if <file>.ocr.txt exists alongside the PDF,
                           read that instead of running docling (fast, no timeout risk).
                           If no sidecar: docling (DEFAULT). OCR is gated: ON only when
                           the PDF has NO readable text layer; OFF when a readable text
                           layer exists.
  .xlsx                 -> openpyxl
  .xls                  -> SKIPPED (old format; needs xlrd) -- documented gap
  other                 -> skipped

DOCLING IS THE DEFAULT FOR ALL RUNTIMES (Rick directive 2026-06-15). docling remains
the engine for EVERY pdf that lacks a sidecar. v2 change (2026-06-16) — OCR GATING:
docling's full-page OCR pass is what pushed large/scanned PDFs past the watcher's
per-file timeout (quarantining 18 docs, 14 of them text-bearing). We now probe for a
text layer first; text-bearing PDFs run docling with do_ocr=False (proven 506pp HP13:
>600s timeout -> 273s, tables intact), and only genuinely scanned PDFs incur the OCR
pass. do_table_structure stays ON for all. pypdf remains an EMERGENCY fallback ONLY
and is LOUD.
v3 change (2026-06-18) — READABILITY CHECK: PDFs with Identity-H / Windows-1252 shift
fonts produce a character-count that passes the text-layer probe but is >40% non-ASCII
garbage. Added _is_readable() check so garbled-encoding PDFs now correctly get OCR.
v4 change (2026-06-19) — Added Coin Database to DEFAULT_FOLDERS.
v5 change (2026-06-20) — CONFIG CONSOLIDATION (GUI Mods_V3 #1): EMBED_MODEL, COLLECTION,
and CHAT_MODEL now come from rag_config.py (single source of truth). CHAT_MODEL is
DEFAULT_MODEL (27B). Retrieval/ingestion logic UNCHANGED from v4 — byte-for-byte.
v6 change (2026-06-22) — SIDECAR-FIRST: if <file>.ocr.txt exists alongside a PDF,
read that instead of invoking docling. Eliminates timeout quarantine for pre-OCR'd
scanned docs. Sidecar is always preferred; docling path unchanged for PDFs without one.

Commands:
  python3 jarvis_rag_poc.py ingest [folder ...]   # default = the 4 canonical folders; RECREATES collection
  python3 jarvis_rag_poc.py ingest-file <path>    # single file, upsert (no recreate) -- used by the watcher
  python3 jarvis_rag_poc.py ask "question"
  python3 jarvis_rag_poc.py count
"""
import sys, os, json, glob, time, uuid, zipfile, subprocess, urllib.request, urllib.error, ssl, re

# rag_config lives in this script's directory; ensure it resolves both when this
# file is run standalone AND when smedley-rag-api.py imports it.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rag_config import EMBED_MODEL, COLLECTION, DEFAULT_MODEL

# ---------- config ----------
LMS         = "http://127.0.0.1:1234/v1"
CHAT_MODEL  = DEFAULT_MODEL                      # from rag_config (27B default)
QDRANT      = "http://192.168.0.25:6333"
QKEY_DOCX   = "/Users/rick/Mounts/Z/DATA/n8n_share/API Keys/qdrant API.docx"
LIB_ROOT    = "/Users/rick/Mounts/RAG_Pool/Library"
DEFAULT_FOLDERS = [
    LIB_ROOT + "/Coin Database",
    LIB_ROOT + "/Electrical Resources",
    LIB_ROOT + "/Network Mappings",
    LIB_ROOT + "/Vendor Data",
]
CHUNK_CHARS = 1000
CHUNK_OVERLAP = 150
EMB_BATCH = 16
TOPK = 6
PDFTOTEXT   = "/opt/homebrew/bin/pdftotext"     # poppler; used only for the fast text-layer probe
TEXT_PROBE_PAGES = 3                            # pages sampled to detect a text layer
TEXT_PROBE_MIN   = 120                          # >= this many chars over the sample => text-bearing
_SSL = ssl.create_default_context(); _SSL.check_hostname = False; _SSL.verify_mode = ssl.CERT_NONE

# ---------- qdrant key ----------
def qdrant_key():
    with zipfile.ZipFile(QKEY_DOCX) as z:
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
    txt = re.sub(r"<[^>]+>", " ", xml)
    m = re.findall(r"[A-Za-z0-9_\-]{30,}", txt)
    if not m:
        raise RuntimeError("no qdrant key found in docx")
    return max(m, key=len)
QKEY = qdrant_key()

# ---------- http helpers ----------
def _req(url, data=None, headers=None, method=None, ctx=None, timeout=120):
    h = {"Content-Type": "application/json"}
    if headers: h.update(headers)
    body = json.dumps(data).encode() if data is not None else None
    r = urllib.request.Request(url, data=body, headers=h, method=method)
    with urllib.request.urlopen(r, context=ctx, timeout=timeout) as resp:
        return json.loads(resp.read().decode())

def qd(path, data=None, method="POST"):
    return _req(QDRANT + path, data=data, headers={"api-key": QKEY}, method=method, ctx=_SSL)

def embed(texts):
    out = []
    for i in range(0, len(texts), EMB_BATCH):
        batch = texts[i:i+EMB_BATCH]
        res = _req(LMS + "/embeddings", {"model": EMBED_MODEL, "input": batch})
        out.extend([d["embedding"] for d in res["data"]])
    return out

def chat(messages):
    res = _req(LMS + "/chat/completions",
               {"model": CHAT_MODEL, "messages": messages, "temperature": 0.2, "stream": False},
               timeout=300)
    return res["choices"][0]["message"]["content"]

# ---------- extraction ----------
def _textutil(path):
    r = subprocess.run(["textutil", "-convert", "txt", "-stdout", path],
                       capture_output=True, text=True, timeout=120)
    return r.stdout or ""

def _is_readable(sample, max_garble=0.40):
    """Return True if sample text is human-readable (not garbled encoding).
    PDFs with Identity-H / Windows-1252 shift fonts produce text that passes
    a character-count check but is >40% non-ASCII garbage. Treat those as
    no-text-layer so OCR runs instead."""
    if not sample:
        return False
    non_ascii = sum(1 for c in sample if ord(c) > 127)
    return (non_ascii / len(sample)) < max_garble

def _pdf_has_text_layer(path):
    """Fast probe: does this PDF carry a READABLE text layer? Drives OCR gating.
    poppler pdftotext on a small page sample; pypdf as a secondary probe. On any
    uncertainty we return False -> docling runs OCR (the safe, correct-but-slow path).
    v3 (2026-06-18): added readability check — garbled Identity-H/Windows-1252 encoded
    fonts produce enough chars to pass the length test but are >40% non-ASCII garbage;
    those PDFs now correctly get OCR instead of silently indexing junk."""
    try:
        r = subprocess.run([PDFTOTEXT, "-f", "1", "-l", str(TEXT_PROBE_PAGES), path, "-"],
                           capture_output=True, text=True, timeout=60)
        sample = (r.stdout or "").strip()
        if len(sample) >= TEXT_PROBE_MIN and _is_readable(sample):
            return True
    except Exception:
        pass
    try:
        from pypdf import PdfReader
        rd = PdfReader(path)
        t = "".join((rd.pages[i].extract_text() or "")
                    for i in range(min(TEXT_PROBE_PAGES, len(rd.pages))))
        sample = t.strip()
        return len(sample) >= TEXT_PROBE_MIN and _is_readable(sample)
    except Exception:
        return False

def _pdf(path):
    # SIDECAR-FIRST (v6): if a pre-OCR'd <file>.ocr.txt exists alongside this PDF,
    # read that instead of invoking docling. Zero timeout risk, full text fidelity.
    # Sidecar is always preferred; docling runs only when no sidecar is present.
    base = os.path.basename(path)
    sidecar = os.path.splitext(path)[0] + ".ocr.txt"
    if os.path.isfile(sidecar):
        try:
            with open(sidecar, "r", errors="ignore") as f:
                text = f.read()
            if text.strip():
                print("      [pdf] sidecar OK  %s  -> %s" % (base, os.path.basename(sidecar)), flush=True)
                return text
            else:
                print("      [pdf] sidecar empty, falling through to docling  %s" % base, flush=True)
        except Exception as e:
            print("      [pdf] sidecar read failed (%s), falling through to docling  %s" % (e, base), flush=True)

    # DOCLING IS THE DEFAULT (engine for every pdf without a sidecar). OCR is GATED:
    # on only when there is no readable text layer. pypdf is an emergency fallback
    # ONLY, and it is LOUD: a missing-docling runtime must never again degrade
    # silently (Rick directive 2026-06-15).
    try:
        from docling.document_converter import DocumentConverter, PdfFormatOption
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        has_text = _pdf_has_text_layer(path)
        opts = PdfPipelineOptions()
        opts.do_ocr = not has_text          # OCR only when there is no readable text layer
        opts.do_table_structure = True      # keep table fidelity on all
        conv = DocumentConverter(format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=opts)})
        md = conv.convert(path).document.export_to_markdown()
        print("      [pdf] docling OK  (ocr=%s)  %s" % ("on" if opts.do_ocr else "off", base), flush=True)
        return md
    except ImportError:
        print("      [pdf] !! DOCLING NOT IMPORTABLE in this runtime (python %s) -- "
              "this runtime is NOT docling-default. FALLING BACK to pypdf (NO OCR): %s"
              % (sys.version.split()[0], base), flush=True)
    except Exception as e:
        print("      [pdf] !! docling FAILED on this file (%s: %s) -- "
              "falling back to pypdf for this file only: %s"
              % (e.__class__.__name__, str(e)[:120], base), flush=True)
    from pypdf import PdfReader
    reader = PdfReader(path)
    return "\n".join((pg.extract_text() or "") for pg in reader.pages)

def _xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append("# Sheet: %s" % ws.title)
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells: parts.append("\t".join(cells))
    return "\n".join(parts)

def extract(path):
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".doc", ".docx", ".rtf", ".html", ".htm"):
            return _textutil(path)
        if ext in (".txt", ".md", ".log", ".csv"):
            with open(path, "r", errors="ignore") as f: return f.read()
        if ext == ".pdf":
            return _pdf(path)
        if ext == ".xlsx":
            return _xlsx(path)
    except Exception as e:
        print("      ! extract fail %s : %s" % (os.path.basename(path), e))
        return ""
    return ""  # unsupported (.xls, .msg, .lnk, .accdb ...)

# ---------- chunking ----------
def chunks_of(text):
    text = re.sub(r"[ \t]+", " ", text).strip()
    if not text: return []
    out, i, n = [], 0, len(text)
    while i < n:
        end = min(i + CHUNK_CHARS, n)
        out.append(text[i:end])
        if end == n: break
        i = end - CHUNK_OVERLAP
    return out

def pid(source, idx):
    return str(uuid.uuid5(uuid.NAMESPACE_URL, "%s#%d" % (source, idx)))

# ---------- collection ----------
def recreate():
    try: qd("/collections/" + COLLECTION, method="DELETE")
    except Exception: pass
    qd("/collections/" + COLLECTION,
       {"vectors": {"size": 768, "distance": "Cosine"}}, method="PUT")

def count():
    r = qd("/collections/%s/points/count" % COLLECTION, {"exact": True})
    return r["result"]["count"]

def delete_source(source):
    qd("/collections/%s/points/delete" % COLLECTION,
       {"filter": {"must": [{"key": "source", "match": {"value": source}}]}})

# ---------- ingest ----------
def list_files(folder):
    return [p for p in glob.glob(os.path.join(folder, "**", "*"), recursive=True)
            if os.path.isfile(p) and "/~" not in p and "/." not in p]

def ingest_one(path, total_holder):
    source = os.path.relpath(path, LIB_ROOT)
    text = extract(path)
    cks = chunks_of(text)
    if not cks: return 0
    vecs = embed(cks)
    pts = [{"id": pid(source, i), "vector": vecs[i],
            "payload": {"source": source, "abspath": path, "chunk": i, "text": cks[i]}}
           for i in range(len(cks))]
    for j in range(0, len(pts), 64):
        qd("/collections/%s/points" % COLLECTION, {"points": pts[j:j+64]}, method="PUT")
    total_holder[0] += len(pts)
    return len(pts)

def ingest(folders, recreate_first=True):
    if recreate_first:
        print("[1/4] (re)creating collection '%s'" % COLLECTION); recreate()
    total = [0]
    allfiles = []
    for f in folders:
        fs = list_files(f)
        print("[corpus] %s -> %d files" % (f, len(fs)))
        allfiles += [(f, p) for p in fs]
    print("[2/4] extracting + embedding %d files ..." % len(allfiles))
    done = 0
    for _, p in allfiles:
        n = ingest_one(p, total)
        done += 1
        if n: print("      [%d/%d] %-55s +%d  (total %d)" % (done, len(allfiles), os.path.basename(p)[:55], n, total[0]))
    print("[3/4] verifying count ...")
    c = count()
    print("[4/4] DONE. collection '%s' holds %d vectors." % (COLLECTION, c))

def ingest_file(path):
    source = os.path.relpath(path, LIB_ROOT)
    print("ingest-file: %s" % source)
    delete_source(source)            # clear old chunks for this file (handles edits)
    total = [0]
    n = ingest_one(path, total)
    print("  upserted %d chunks. collection holds %d." % (n, count()))

# ---------- ask ----------
def ask(question):
    qv = embed([question])[0]
    res = qd("/collections/%s/points/search" % COLLECTION,
             {"vector": qv, "limit": TOPK, "with_payload": True})
    hits = res["result"]
    if not hits:
        print("No matches. Is the collection populated? (count = %d)" % count()); return
    ctx = "\n\n".join("[%d] (%s)\n%s" % (i+1, h["payload"]["source"], h["payload"]["text"])
                      for i, h in enumerate(hits))
    sys_p = ("/no_think You are a precise engineering assistant. Answer ONLY from the "
             "provided context. Cite the source filename in brackets. If the context does "
             "not contain the answer, say so plainly.")
    ans = chat([{"role": "system", "content": sys_p},
                {"role": "user", "content": "Context:\n%s\n\nQuestion: %s" % (ctx, question)}])
    ans = re.sub(r"<think>.*?</think>", "", ans, flags=re.S).strip()
    print("\n=== ANSWER ===\n" + ans)
    print("\n=== RETRIEVED FROM ===")
    seen = set()
    for h in hits:
        s = h["payload"]["source"]
        tag = "%s (score %.3f)" % (s, h["score"])
        if s not in seen:
            print("  - " + tag); seen.add(s)

# ---------- main ----------
if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "ingest":
        folders = sys.argv[2:] or DEFAULT_FOLDERS
        ingest(folders, recreate_first=True)
    elif cmd == "ingest-file":
        ingest_file(sys.argv[2])
    elif cmd == "ask":
        ask(sys.argv[2])
    elif cmd == "count":
        print("collection '%s' holds %d vectors." % (COLLECTION, count()))
    else:
        print(__doc__)
